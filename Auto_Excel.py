import openpyxl
wb = openpyxl.load_workbook('files.xlsx')
print(wb.get_sheet_names())
sheet1 = wb.get_sheet_by_name('sheet name')
wb2 = openpyxl.load_workbook('file2.xlsx')
print(wb2.get_sheet_names())
sheet2 = wb2.get_sheet_by_name('sheetname')
for i in range(2,sheet1.max_row+1):
    for j in range(2,sheet2.max_row+1):
        if  sheet1.cell(row=i,column=4).value == sheet2.cell(row=j,column=3).value:
            sheet2.cell(row=j,column=6).value = sheet1.cell(row=i,column=16).value
            #sheet2.cell(row=j,column=6).value = sheet1.cell(row=i,column=25).value
            print(sheet1.cell(row=j,column=16).value)
            break
        else:
            continue
        

wb2.save('files.xlsx')
